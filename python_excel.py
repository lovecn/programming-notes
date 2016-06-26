from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active              #第一个表
ws1.title = "range names"    #第一个表命名
#遍历第一个表的1到39行，每行赋值从1到599.https://segmentfault.com/a/1190000005144821
for row in range(1,40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi") # 创建第二个表
ws2['F5'] = 3.14     #为第二个表的F5单元格赋值为3.14
ws3 = wb.create_sheet(title="Data")  #创建第三个表
 /* 下面遍历第三个表的10到19行，27到53列，并对每一行的单元格赋一个当前列名的名字如下图 */
for row in range(10,20):
    for col in range(27,54):
        _=ws3.cell(column=col,row=row,value="%s" % get_column_letter(col)) #_当作一个普通的变量，一般表示后边不再使用
wb.save(filename=dest_filename) #保存
